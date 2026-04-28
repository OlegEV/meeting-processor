#!/usr/bin/env python3
"""
Рабочая версия веб-приложения для обработки встреч
Полностью протестирована и исправлена
"""

import os
import sys
import json
import uuid
import threading
from pathlib import Path
from typing import Dict, Optional, Any
import logging
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor

# Загружаем переменные окружения из .env файла
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    # python-dotenv не установлен, переменные окружения должны быть установлены системно
    pass

# Веб-фреймворк
try:
    from flask import Flask, request, render_template_string, jsonify, send_file, redirect, url_for, flash, session
    from werkzeug.utils import secure_filename as werkzeug_secure_filename
    from werkzeug.exceptions import RequestEntityTooLarge
except ImportError:
    print("❌ Установите Flask: pip install Flask")
    sys.exit(1)

# Наши модули
try:
    from meeting_processor import MeetingProcessor
    from config_loader import ConfigLoader
    from web_templates import WebTemplates
    from auth import create_auth_system, require_auth, get_current_user_id, get_current_user, is_authenticated
    from database import create_database_manager
    # Confluence модули (опциональные)
    confluence_available = True
    try:
        from confluence_client import ConfluenceServerClient, ConfluenceConfig, ConfluencePublicationService
        from confluence_encryption import create_token_manager
    except ImportError as confluence_error:
        confluence_available = False
        logger = logging.getLogger(__name__)
        logger.warning(f"Confluence модули недоступны: {confluence_error}")
except ImportError as e:
    print(f"❌ Ошибка импорта модулей: {e}")
    sys.exit(1)

# Настройка логирования
def setup_logging(log_level: str = "DEBUG", log_file: str = "web_app.log"):
    """Настраивает систему логирования"""
    from logging.handlers import RotatingFileHandler
    
    level = getattr(logging, log_level.upper(), logging.INFO)
    
    # Создаем папку для логов
    os.makedirs("logs", exist_ok=True)
    log_path = os.path.join("logs", log_file)
    
    # Очищаем существующие обработчики
    root_logger = logging.getLogger()
    for handler in root_logger.handlers[:]:
        root_logger.removeHandler(handler)
    
    handlers = [
        RotatingFileHandler(log_path, maxBytes=100*1024*1024, backupCount=3, encoding='utf-8'),
        logging.StreamHandler()
    ]
    
    logging.basicConfig(
        level=level,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=handlers,
        force=True  # Принудительно переопределяем конфигурацию
    )
    
    # Настраиваем логгеры для всех модулей приложения
    app_logger = logging.getLogger(__name__)
    app_logger.setLevel(level)
    
    # Настраиваем логгер для Flask
    flask_logger = logging.getLogger('werkzeug')
    flask_logger.setLevel(logging.WARNING)  # Уменьшаем уровень для werkzeug
    
    return app_logger

logger = setup_logging()

def secure_filename_unicode(filename: str) -> str:
    """
    Безопасная обработка имени файла с поддержкой русских символов
    Сохраняет расширение файла и заменяет только опасные символы
    """
    if not filename:
        return filename
    
    # Разделяем имя файла и расширение
    name_part, ext_part = os.path.splitext(filename)
    
    # Список опасных символов для замены
    dangerous_chars = ['/', '\\', ':', '*', '?', '"', '<', '>', '|', '\0']
    
    # Заменяем опасные символы на подчеркивания
    safe_name = name_part
    for char in dangerous_chars:
        safe_name = safe_name.replace(char, '_')
    
    # Убираем множественные подчеркивания и пробелы в начале/конце
    safe_name = '_'.join(safe_name.split())
    safe_name = safe_name.strip('._')
    
    # Если имя стало пустым, используем fallback
    if not safe_name:
        safe_name = 'file'
    
    # Возвращаем безопасное имя с оригинальным расширением
    result = safe_name + ext_part.lower()
    
    logger.debug(f"Filename sanitization: '{filename}' -> '{result}'")
    return result

class WorkingMeetingWebApp:
    """Рабочая версия веб-приложения для обработки встреч"""
    
    def __init__(self, config_file: str = "config.json"):
        self.app = Flask(__name__)
        self.app.secret_key = os.urandom(24)
        
        # Загружаем конфигурацию
        self.config = ConfigLoader.load_config(config_file)
        if not self.config:
            raise Exception(f"Не удалось загрузить конфигурацию из {config_file}")
        
        # Загружаем API ключи из переменных окружения
        self.api_keys_data = ConfigLoader.load_api_keys()
        
        # Проверяем API ключи
        deepgram_valid, claude_valid, self.deepgram_key, self.claude_key = ConfigLoader.validate_api_keys(self.api_keys_data)
        
        if not deepgram_valid or not claude_valid:
            raise Exception("❌ API ключи не настроены")
        
        # Инициализируем систему аутентификации
        self.token_validator, self.user_manager, self.auth_middleware, self.auth_teardown = create_auth_system(self.config)
        
        # Инициализируем базу данных
        self.db_manager = create_database_manager(self.config)
        
        # Связываем user_manager с db_manager
        self.user_manager.set_db_manager(self.db_manager)
        
        # Инициализируем Confluence (если доступен и настроен)
        self.confluence_client = None
        self.confluence_service = None
        self.token_manager = None
        self._init_confluence()
        
        # Настройки приложения
        self.setup_app_config()
        
        # Пул потоков для обработки файлов
        self.executor = ThreadPoolExecutor(max_workers=3, thread_name_prefix="FileProcessor")
        
        # Создаем необходимые директории
        self.upload_folder = Path("web_uploads")
        self.output_folder = Path("web_output")
        self.upload_folder.mkdir(exist_ok=True)
        self.output_folder.mkdir(exist_ok=True)
        
        # Инициализируем шаблоны
        self.templates = WebTemplates()
        
        # Настраиваем middleware
        self.setup_middleware()
        
        # Настраиваем маршруты
        self.setup_routes()
        
        logger.info("🌐 Веб-приложение инициализировано с аутентификацией и базой данных")
    
    def setup_app_config(self):
        """Настраивает конфигурацию Flask"""
        # Максимальный размер файла
        max_size_mb = self.config.get("settings", {}).get("max_file_size_mb", 200)
        self.app.config['MAX_CONTENT_LENGTH'] = max_size_mb * 1024 * 1024
        
        # Поддерживаемые форматы файлов
        self.allowed_extensions = {'mp3', 'wav', 'flac', 'aac', 'm4a', 'ogg', 'opus', 'mp4', 'avi', 'mov', 'mkv', 'wmv', 'webm'}
        
        # Пытаемся загрузить из конфига
        formats = self.config.get("supported_formats", {})
        if formats:
            self.allowed_extensions = set()
            for format_list in formats.values():
                clean_extensions = [ext.lstrip('.').lower() for ext in format_list]
                self.allowed_extensions.update(clean_extensions)
        
        # Настройки обработки
        self.processing_settings = self.config.get("settings", {})
        
        # Создаем строки для HTML
        self.allowed_extensions_list = sorted(list(self.allowed_extensions))
        self.accept_string = ','.join([f'.{ext}' for ext in self.allowed_extensions_list])
        self.formats_display = ', '.join([ext.upper() for ext in self.allowed_extensions_list])
        
        logger.info(f"Максимальный размер файла: {max_size_mb} МБ")
        logger.info(f"Поддерживаемые форматы: {self.formats_display}")
    
    def setup_middleware(self):
        """Настраивает middleware для аутентификации"""
        # Middleware для аутентификации
        self.app.before_request(self.auth_middleware)
        
        # Teardown для очистки контекста
        self.app.teardown_appcontext(self.auth_teardown)
        
        logger.info("Middleware аутентификации настроен")
    
    def _init_confluence(self):
        """Инициализирует Confluence интеграцию"""
        try:
            if not confluence_available:
                logger.info("Confluence модули недоступны, пропускаем инициализацию")
                return
            
            confluence_config = self.config.get('confluence', {})
            if not confluence_config.get('enabled', False):
                logger.info("Confluence интеграция отключена в конфигурации")
                return
            
            # Проверяем обязательные настройки
            required_settings = ['base_url', 'space_key']
            missing_settings = [setting for setting in required_settings
                              if not confluence_config.get(setting)]
            
            if missing_settings:
                logger.warning(f"Confluence: отсутствуют обязательные настройки: {missing_settings}")
                return
            
            # Инициализируем менеджер токенов
            try:
                self.token_manager = create_token_manager("confluence_tokens.json")
                logger.info("Менеджер токенов Confluence инициализирован")
            except Exception as e:
                logger.error(f"Ошибка инициализации менеджера токенов: {e}")
                return
            
            # Получаем токен API из переменных окружения или конфигурации
            api_token = os.getenv('CONFLUENCE_API_TOKEN')
            
            if not api_token and confluence_config.get('encrypted_token'):
                try:
                    # Пытаемся получить расшифрованный токен
                    api_token = self.token_manager.get_token(
                        confluence_config['username'],
                        confluence_config.get('encryption_key', '')
                    )
                except Exception as e:
                    logger.warning(f"Не удалось расшифровать токен: {e}")
            
            if not api_token and confluence_config.get('api_token'):
                api_token = confluence_config['api_token']
            
            if not api_token:
                logger.warning("Confluence: API токен не найден в переменных окружения или конфигурации")
                return
            
            # Создаем конфигурацию Confluence
            try:
                config = ConfluenceConfig(
                    base_url=confluence_config['base_url'],
                    api_token=api_token,
                    space_key=confluence_config['space_key'],
                    username=confluence_config.get('username'),  # Опциональное поле для совместимости
                    parent_page_id=confluence_config.get('parent_page_id'),
                    timeout=confluence_config.get('timeout', 30),
                    max_retries=confluence_config.get('max_retries', 3),
                    retry_delay=confluence_config.get('retry_delay', 1.0)
                )
                
                # Создаем клиент Confluence
                self.confluence_client = ConfluenceServerClient(config)
                
                # Создаем сервис публикации
                self.confluence_service = ConfluencePublicationService(
                    self.confluence_client,
                    self.db_manager
                )
                
                logger.info("Confluence интеграция успешно инициализирована")
                
            except Exception as e:
                logger.error(f"Ошибка создания Confluence клиента: {e}")
                self.confluence_client = None
                self.confluence_service = None
        
        except Exception as e:
            logger.error(f"Ошибка инициализации Confluence: {e}")
            self.confluence_client = None
            self.confluence_service = None
    
    def allowed_file(self, filename: str) -> bool:
        """Проверяет, разрешен ли файл для загрузки"""
        if not filename or '.' not in filename:
            logger.warning(f"Файл без расширения: '{filename}'")
            return False
        
        try:
            file_ext = filename.rsplit('.', 1)[1].lower()
            is_allowed = file_ext in self.allowed_extensions
            
            if not is_allowed:
                logger.warning(f"Неподдерживаемое расширение: '{file_ext}' в файле '{filename}'")
            else:
                logger.debug(f"Файл разрешен: '{filename}' (расширение: '{file_ext}')")
            
            return is_allowed
        except Exception as e:
            logger.error(f"Ошибка проверки расширения файла '{filename}': {e}")
            return False
    
    def get_available_templates(self) -> Dict[str, str]:
        """Возвращает доступные шаблоны.

        Источник истины — templates_config.json (template_descriptions).
        Тот же файл читают meeting_templates.py и template_manager.py,
        благодаря чему веб-UI, CLI-утилиты и core-обработчик видят одинаковый
        список шаблонов.
        """
        templates_config_path = self.config.get("paths", {}).get(
            "templates_config", "templates_config.json"
        )
        try:
            with open(templates_config_path, "r", encoding="utf-8") as f:
                templates_config = json.load(f)
            descriptions = templates_config.get("template_descriptions", {})
            if descriptions:
                return descriptions
        except (FileNotFoundError, json.JSONDecodeError) as e:
            logger.warning(
                f"Не удалось прочитать template_descriptions из {templates_config_path}: {e}"
            )

        # Фоллбэк, если файл недоступен или пуст
        return {
            "standard": "Универсальный шаблон для любых встреч",
            "business": "Официальный протокол для деловых встреч",
            "project": "Фокус на управлении проектами и задачами",
            "standup": "Краткий формат для ежедневных стендапов",
            "interview": "Структурированный отчет об интервью",
            "auto": "Автоматическое определение типа встречи"
        }

    def get_available_models(self) -> Dict[str, str]:
        """Возвращает доступные модели OpenRouter (id -> человеко-читаемое описание)"""
        return self.config.get("available_models", {
            "anthropic/claude-sonnet-4.6": "Claude Sonnet 4.6",
            "anthropic/claude-haiku-4.5": "Claude Haiku 4.5",
            "moonshotai/kimi-k2.6": "Kimi K2.6"
        })

    def get_default_model(self) -> str:
        """Возвращает модель по умолчанию (из settings, с fallback на первую из available_models)"""
        default = self.processing_settings.get('claude_model')
        available = self.get_available_models()
        if default and default in available:
            return default
        return next(iter(available), 'anthropic/claude-sonnet-4.6')

    def resolve_model(self, requested: Optional[str]) -> str:
        """Валидирует выбранную пользователем модель, возвращает её или модель по умолчанию"""
        available = self.get_available_models()
        if requested and requested in available:
            return requested
        return self.get_default_model()

    def get_job_model(self, job: Dict[str, Any]) -> str:
        """Достаёт сохранённую в metadata задачи модель, либо возвращает дефолт"""
        metadata = job.get('metadata') if job else None
        if isinstance(metadata, str):
            try:
                import json as _json
                metadata = _json.loads(metadata) if metadata else {}
            except (ValueError, TypeError):
                metadata = {}
        if isinstance(metadata, dict):
            model = metadata.get('model')
            if model:
                return self.resolve_model(model)
        return self.get_default_model()
    
    def update_job_status(self, job_id: str, **kwargs):
        """Безопасно обновляет статус задачи в базе данных"""
        self.update_job_in_db(job_id, kwargs)
    
    def get_job_status(self, job_id: str) -> Optional[Dict]:
        """Безопасно получает статус задачи из базы данных"""
        try:
            # Получаем текущего пользователя
            user_id = get_current_user_id()
            if not user_id:
                logger.warning("Попытка получить статус задачи без аутентификации")
                return None
            
            # Получаем задачу из базы данных с проверкой доступа
            job_data = self.db_manager.get_job_by_id(job_id, user_id)
            return job_data
            
        except Exception as e:
            logger.error(f"Ошибка получения статуса задачи {job_id}: {e}")
            return None
    
    def create_job_in_db(self, job_data: Dict[str, Any]) -> bool:
        """Создает задачу в базе данных"""
        try:
            user_id = get_current_user_id()
            if not user_id:
                logger.error("Попытка создать задачу без аутентификации")
                return False
            
            # Добавляем user_id к данным задачи
            job_data['user_id'] = user_id
            
            # Создаем задачу в базе данных
            self.db_manager.create_job(job_data)
            logger.info(f"Задача {job_data['job_id']} создана в базе данных для пользователя {user_id}")
            return True
            
        except Exception as e:
            logger.error(f"Ошибка создания задачи в базе данных: {e}")
            return False
    
    def update_job_in_db(self, job_id: str, update_data: Dict[str, Any]) -> bool:
        """Обновляет задачу в базе данных"""
        try:
            user_id = get_current_user_id()
            if not user_id:
                logger.error("Попытка обновить задачу без аутентификации")
                return False
            
            # Обновляем задачу в базе данных с проверкой доступа
            success = self.db_manager.update_job(job_id, update_data, user_id)
            if success:
                logger.debug(f"Задача {job_id} обновлена в базе данных")
            return success
            
        except Exception as e:
            logger.error(f"Ошибка обновления задачи {job_id}: {e}")
            return False
    
    def get_user_output_dir(self, user_id: str) -> Path:
        """Получает директорию для файлов пользователя"""
        user_files_config = self.config.get('user_files', {})
        base_path = user_files_config.get('base_path', 'web_output')
        
        if user_files_config.get('structure') == 'user_based':
            return Path(base_path) / user_id
        else:
            return Path(base_path)
    
    def ensure_user_exists(self) -> Optional[Dict[str, Any]]:
        """Обеспечивает существование пользователя в базе данных"""
        try:
            user_info = get_current_user()
            if not user_info:
                return None
            
            # Создаем или обновляем пользователя в базе данных
            db_user = self.user_manager.ensure_user_exists(user_info)
            return db_user
            
        except Exception as e:
            logger.error(f"Ошибка при работе с пользователем: {e}")
            return None
    
    def extract_confluence_metadata(self, base_page_url: str, api_token: str = None, timeout: int = 30) -> Dict[str, Optional[str]]:
        """
        Универсальная функция для извлечения метаданных Confluence страницы
        
        Args:
            base_page_url: URL страницы Confluence
            api_token: API токен для аутентификации (опционально)
            timeout: Таймаут для HTTP запросов
            
        Returns:
            Dict с извлеченными метаданными:
            {
                'page_id': str | None,
                'space_key': str | None,
                'page_title': str | None,
                'base_url': str | None,
                'extraction_method': str  # Метод, которым были извлечены данные
            }
        """
        import re
        import requests
        import json
        from urllib.parse import urlparse, parse_qs, unquote
        
        result = {
            'page_id': None,
            'space_key': None,
            'page_title': None,
            'base_url': None,
            'extraction_method': 'none'
        }
        
        if not base_page_url:
            logger.warning("🔍 extract_confluence_metadata: Пустой URL")
            return result
        
        logger.info(f"🔍 extract_confluence_metadata: Анализ URL: {base_page_url}")
        
        try:
            # Извлекаем базовый URL
            parsed_url = urlparse(base_page_url)
            result['base_url'] = f"{parsed_url.scheme}://{parsed_url.netloc}"
            
            # 1. Попытка извлечения из URL паттернов
            logger.info("🔍 Попытка 1: Извлечение из URL паттернов")
            
            # Confluence Server формат 1: /pages/viewpage.action?pageId=123456
            server_pattern1 = r'/pages/viewpage\.action\?pageId=(\d+)'
            match = re.search(server_pattern1, base_page_url)
            if match:
                result['page_id'] = match.group(1)
                result['extraction_method'] = 'url_viewpage'
                logger.info(f"🔍 Найден page_id из viewpage URL: {result['page_id']}")
            
            # Confluence Server формат 2: /display/SPACE/PAGE
            server_pattern2 = r'/display/([^/]+)/(.+?)(?:\?|$)'
            match = re.search(server_pattern2, base_page_url)
            if match:
                result['space_key'] = match.group(1)
                page_slug = unquote(match.group(2))  # Декодируем URL-encoded символы
                result['extraction_method'] = 'url_display'
                logger.info(f"🔍 Найден space_key из display URL: {result['space_key']}")
                logger.info(f"🔍 Найден page slug: {page_slug}")
                
                # Попытка получить page_id через Confluence REST API
                logger.info(f"🔍 DEBUG REST API: api_token={'есть' if api_token else 'НЕТ'}, space_key={result['space_key']}, base_url={result['base_url']}")
                
                if api_token and result['space_key'] and result['base_url']:
                    try:
                        # Заменяем + на пробелы в заголовке
                        page_title_search = page_slug.replace('+', ' ')
                        
                        logger.info(f"🔍 Попытка получить page_id через Confluence REST API для '{page_title_search}' в пространстве '{result['space_key']}'")
                        
                        # Используем Confluence REST API для поиска страницы по заголовку
                        api_url = f"{result['base_url']}/rest/api/content"
                        params = {
                            'type': 'page',
                            'spaceKey': result['space_key'],
                            'title': page_title_search,
                            'expand': 'version,space'
                        }
                        
                        headers = {
                            'Authorization': f'Bearer {api_token}',
                            'Accept': 'application/json'
                        }
                        
                        response = requests.get(api_url, params=params, headers=headers, timeout=timeout)
                        
                        if response.status_code == 200:
                            data = response.json()
                            
                            if data.get('results') and len(data['results']) > 0:
                                page_info = data['results'][0]
                                result['page_id'] = str(page_info['id'])
                                result['page_title'] = page_info.get('title', page_title_search)
                                result['extraction_method'] = 'rest_api'
                                logger.info(f"🔍 ✅ Получен page_id через REST API: {result['page_id']}")
                                logger.info(f"🔍 ✅ Получен page_title через REST API: {result['page_title']}")
                            else:
                                logger.warning(f"🔍 REST API не нашел страницу '{page_title_search}' в пространстве '{result['space_key']}'")
                        else:
                            logger.warning(f"🔍 REST API вернул статус {response.status_code}: {response.text[:200]}")
                    
                    except requests.RequestException as api_error:
                        logger.warning(f"🔍 Ошибка при использовании Confluence REST API: {api_error}")
                        # Продолжаем с другими методами извлечения
                    except Exception as api_error:
                        logger.warning(f"🔍 Неожиданная ошибка при использовании REST API: {api_error}")
                        # Продолжаем с другими методами извлечения
            
            # Confluence Cloud формат: /wiki/spaces/SPACE/pages/123456/PAGE
            cloud_pattern = r'/wiki/spaces/([^/]+)/pages/(\d+)/(.+?)(?:\?|$)'
            match = re.search(cloud_pattern, base_page_url)
            if match:
                result['space_key'] = match.group(1)
                result['page_id'] = match.group(2)
                page_slug = match.group(3)
                result['extraction_method'] = 'url_cloud'
                logger.info(f"🔍 Найден space_key из cloud URL: {result['space_key']}")
                logger.info(f"🔍 Найден page_id из cloud URL: {result['page_id']}")
                logger.info(f"🔍 Найден page slug: {page_slug}")
            
            # 2. Попытка извлечения из HTML метаданных страницы
            if not result['page_id'] or not result['space_key']:
                logger.info("🔍 Попытка 2: Извлечение из HTML метаданных")
                
                try:
                    session = requests.Session()
                    headers = {
                        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                    }
                    
                    if api_token:
                        headers['Authorization'] = f'Bearer {api_token}'
                    
                    session.headers.update(headers)
                    
                    response = session.get(base_page_url, timeout=timeout)
                    if response.status_code == 200:
                        html_content = response.text
                        
                        # Паттерны для поиска page ID в HTML
                        page_id_patterns = [
                            r'name="ajs-page-id"\s+content="(\d+)"',  # Confluence Server
                            r'content="(\d+)"\s+name="ajs-page-id"',  # Обратный порядок
                            r'"pageId":"(\d+)"',                      # JSON в скриптах
                            r'"pageId":(\d+)',                        # JSON без кавычек
                            r'pageId["\']?\s*[:=]\s*["\']?(\d+)',     # Различные форматы присваивания
                            r'data-page-id="(\d+)"',                 # Data атрибуты
                            r'pageId=(\d+)',                          # URL параметры в скриптах
                            r'page-id["\']?\s*[:=]\s*["\']?(\d+)',    # Альтернативные названия
                        ]
                        
                        # Паттерны для поиска space key в HTML
                        space_key_patterns = [
                            r'name="ajs-space-key"\s+content="([^"]+)"',  # Confluence Server
                            r'content="([^"]+)"\s+name="ajs-space-key"',  # Обратный порядок
                            r'"spaceKey":"([^"]+)"',                      # JSON в скриптах
                            r'"spaceKey":([^,}\]]+)',                     # JSON без кавычек
                            r'spaceKey["\']?\s*[:=]\s*["\']?([^"\'}\],\s]+)', # Различные форматы
                            r'data-space-key="([^"]+)"',                 # Data атрибуты
                            r'space-key["\']?\s*[:=]\s*["\']?([^"\'}\],\s]+)', # Альтернативные названия
                        ]
                        
                        # Паттерны для поиска заголовка страницы
                        title_patterns = [
                            r'<title>([^<]+)</title>',                    # HTML title
                            r'name="ajs-page-title"\s+content="([^"]+)"', # Confluence meta
                            r'content="([^"]+)"\s+name="ajs-page-title"', # Обратный порядок
                            r'"pageTitle":"([^"]+)"',                     # JSON в скриптах
                            r'data-page-title="([^"]+)"',                # Data атрибуты
                        ]
                        
                        # Извлекаем page ID
                        if not result['page_id']:
                            for pattern in page_id_patterns:
                                match = re.search(pattern, html_content, re.IGNORECASE)
                                if match:
                                    result['page_id'] = match.group(1)
                                    result['extraction_method'] = 'html_metadata'
                                    logger.info(f"🔍 Найден page_id из HTML: {result['page_id']} (паттерн: {pattern})")
                                    break
                        
                        # Извлекаем space key
                        if not result['space_key']:
                            for pattern in space_key_patterns:
                                match = re.search(pattern, html_content, re.IGNORECASE)
                                if match:
                                    space_key_candidate = match.group(1).strip('"\'')
                                    # Проверяем, что это валидный space key (не содержит пробелов и спецсимволов)
                                    if re.match(r'^[A-Z0-9_~-]+$', space_key_candidate, re.IGNORECASE):
                                        result['space_key'] = space_key_candidate
                                        result['extraction_method'] = 'html_metadata'
                                        logger.info(f"🔍 Найден space_key из HTML: {result['space_key']} (паттерн: {pattern})")
                                        break
                        
                        # Извлекаем заголовок страницы
                        if not result['page_title']:
                            for pattern in title_patterns:
                                match = re.search(pattern, html_content, re.IGNORECASE)
                                if match:
                                    title_candidate = match.group(1).strip()
                                    # Очищаем заголовок от лишних частей (например, "- Confluence")
                                    title_candidate = re.sub(r'\s*-\s*Confluence.*$', '', title_candidate)
                                    if title_candidate and len(title_candidate) > 0:
                                        result['page_title'] = title_candidate
                                        logger.info(f"🔍 Найден page_title из HTML: {result['page_title']} (паттерн: {pattern})")
                                        break
                        
                        # 3. Попытка извлечения из JSON-LD данных
                        if not result['page_id'] or not result['space_key']:
                            logger.info("🔍 Попытка 3: Извлечение из JSON-LD данных")
                            
                            # Ищем JSON-LD блоки
                            json_ld_pattern = r'<script[^>]*type=["\']application/ld\+json["\'][^>]*>(.*?)</script>'
                            json_ld_matches = re.findall(json_ld_pattern, html_content, re.DOTALL | re.IGNORECASE)
                            
                            for json_ld_content in json_ld_matches:
                                try:
                                    json_data = json.loads(json_ld_content.strip())
                                    
                                    # Ищем данные в JSON-LD структуре
                                    if isinstance(json_data, dict):
                                        # Проверяем различные поля
                                        if 'identifier' in json_data and not result['page_id']:
                                            identifier = str(json_data['identifier'])
                                            if identifier.isdigit():
                                                result['page_id'] = identifier
                                                result['extraction_method'] = 'json_ld'
                                                logger.info(f"🔍 Найден page_id из JSON-LD: {result['page_id']}")
                                        
                                        if 'name' in json_data and not result['page_title']:
                                            result['page_title'] = str(json_data['name'])
                                            logger.info(f"🔍 Найден page_title из JSON-LD: {result['page_title']}")
                                        
                                        # Ищем space key в различных полях
                                        for field in ['spaceKey', 'space', 'category']:
                                            if field in json_data and not result['space_key']:
                                                space_candidate = str(json_data[field])
                                                if re.match(r'^[A-Z0-9_~-]+$', space_candidate, re.IGNORECASE):
                                                    result['space_key'] = space_candidate
                                                    result['extraction_method'] = 'json_ld'
                                                    logger.info(f"🔍 Найден space_key из JSON-LD: {result['space_key']}")
                                                    break
                                
                                except json.JSONDecodeError:
                                    continue
                        
                        # 4. Попытка извлечения из JavaScript переменных
                        if not result['page_id'] or not result['space_key']:
                            logger.info("🔍 Попытка 4: Извлечение из JavaScript переменных")
                            
                            # Ищем JavaScript объекты с данными
                            js_patterns = [
                                r'AJS\.params\s*=\s*({[^}]+})',
                                r'window\.confluenceData\s*=\s*({[^}]+})',
                                r'var\s+pageData\s*=\s*({[^}]+})',
                                r'window\.pageData\s*=\s*({[^}]+})',
                            ]
                            
                            for js_pattern in js_patterns:
                                matches = re.findall(js_pattern, html_content, re.IGNORECASE)
                                for match in matches:
                                    try:
                                        # Пытаемся извлечь данные из JavaScript объекта
                                        js_data = match
                                        
                                        # Ищем pageId
                                        if not result['page_id']:
                                            page_id_match = re.search(r'["\']?pageId["\']?\s*:\s*["\']?(\d+)', js_data)
                                            if page_id_match:
                                                result['page_id'] = page_id_match.group(1)
                                                result['extraction_method'] = 'javascript'
                                                logger.info(f"🔍 Найден page_id из JavaScript: {result['page_id']}")
                                        
                                        # Ищем spaceKey
                                        if not result['space_key']:
                                            space_key_match = re.search(r'["\']?spaceKey["\']?\s*:\s*["\']([^"\']+)', js_data)
                                            if space_key_match:
                                                space_candidate = space_key_match.group(1)
                                                if re.match(r'^[A-Z0-9_~-]+$', space_candidate, re.IGNORECASE):
                                                    result['space_key'] = space_candidate
                                                    result['extraction_method'] = 'javascript'
                                                    logger.info(f"🔍 Найден space_key из JavaScript: {result['space_key']}")
                                    
                                    except Exception:
                                        continue
                    
                    else:
                        logger.warning(f"🔍 HTTP {response.status_code} при доступе к {base_page_url}")
                
                except requests.RequestException as e:
                    logger.warning(f"🔍 Ошибка HTTP запроса к {base_page_url}: {e}")
                except Exception as e:
                    logger.warning(f"🔍 Ошибка извлечения метаданных из HTML: {e}")
            
            # Логируем итоговый результат
            logger.info(f"🔍 Результат извлечения метаданных:")
            logger.info(f"   page_id: {result['page_id']}")
            logger.info(f"   space_key: {result['space_key']}")
            logger.info(f"   page_title: {result['page_title']}")
            logger.info(f"   base_url: {result['base_url']}")
            logger.info(f"   extraction_method: {result['extraction_method']}")
            
            return result
            
        except Exception as e:
            logger.error(f"🔍 Ошибка в extract_confluence_metadata: {e}")
            result['extraction_method'] = 'error'
            return result
    
    def setup_routes(self):
        """Настраивает маршруты приложения"""
        
        @self.app.route('/health')
        def health_check():
            """Health check для мониторинга"""
            try:
                # Проверяем состояние базы данных
                db_info = self.db_manager.get_database_info()
                
                return jsonify({
                    'status': 'healthy',
                    'timestamp': datetime.utcnow().isoformat(),
                    'version': '1.0.0',
                    'database': {
                        'users_count': db_info.get('users_count', 0),
                        'jobs_count': db_info.get('jobs_count', 0),
                        'db_size_mb': db_info.get('db_size_mb', 0)
                    },
                    'auth': {
                        'enabled': True,
                        'token_header': self.config.get('auth', {}).get('token_header', 'X-Identity-Token')
                    },
                    'confluence': {
                        'available': confluence_available,
                        'enabled': self.config.get('confluence', {}).get('enabled', False),
                        'configured': self.confluence_client is not None
                    }
                })
            except Exception as e:
                logger.error(f"Ошибка health check: {e}")
                return jsonify({
                    'status': 'unhealthy',
                    'timestamp': datetime.utcnow().isoformat(),
                    'error': str(e)
                }), 500
            
        @self.app.route('/')
        @require_auth(redirect_on_failure=False)
        def index():
            """Главная страница"""
            # Обеспечиваем существование пользователя в базе данных
            user = self.ensure_user_exists()
            if not user:
                return jsonify({'error': 'User authentication failed'}), 401
            
            templates = self.get_available_templates()
            available_models = self.get_available_models()
            default_model = self.get_default_model()
            max_size_mb = self.app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024)

            # Получаем информацию о пользователе для отображения
            user_info = get_current_user()
            user_name = self.user_manager.get_user_display_name(user_info) if user_info else "Unknown User"

            return render_template_string(
                self.templates.get_index_template(),
                templates=templates,
                available_models=available_models,
                default_model=default_model,
                max_size_mb=max_size_mb,
                accept_string=self.accept_string,
                formats_display=self.formats_display,
                user_name=user_name,
                user_id=get_current_user_id()
            )
        
        @self.app.route('/upload', methods=['POST'])
        @require_auth()
        def upload_file():
            """Обработка загрузки файла"""
            try:
                # Обеспечиваем существование пользователя
                user = self.ensure_user_exists()
                if not user:
                    flash('Ошибка аутентификации пользователя', 'error')
                    return redirect(url_for('index'))
                
                user_id = get_current_user_id()
                
                if 'file' not in request.files:
                    flash('Файл не выбран', 'error')
                    return redirect(url_for('index'))
                
                file = request.files['file']
                template_type = request.form.get('template', 'standard')
                selected_model = self.resolve_model(request.form.get('model'))

                logger.info(f"📤 Получен файл для загрузки: '{file.filename}' (пользователь: {user_id}, модель: {selected_model})")
                
                if file.filename == '':
                    flash('Файл не выбран', 'error')
                    return redirect(url_for('index'))
                
                if not self.allowed_file(file.filename):
                    flash(f'Неподдерживаемый формат файла. Разрешены: {self.formats_display}', 'error')
                    return redirect(url_for('index'))
                
                # Создаем уникальный ID для задачи
                job_id = str(uuid.uuid4())
                
                # Сохраняем файл в пользовательскую директорию
                original_filename = file.filename
                filename = secure_filename_unicode(file.filename)
                user_upload_dir = self.upload_folder / user_id
                user_upload_dir.mkdir(exist_ok=True)
                file_path = user_upload_dir / f"{job_id}_{filename}"
                
                logger.info(f"📝 Обработка имени файла:")
                logger.info(f"   Оригинальное имя: '{original_filename}'")
                logger.info(f"   Безопасное имя: '{filename}'")
                logger.info(f"   Полный путь: '{file_path}'")
                
                file.save(str(file_path))
                
                logger.info(f"📁 Файл загружен: {filename} (ID: {job_id}, пользователь: {user_id}, шаблон: {template_type})")
                
                # Создаем задачу в базе данных
                job_data = {
                    'job_id': job_id,
                    'user_id': user_id,
                    'filename': filename,
                    'template': template_type,
                    'status': 'uploaded',
                    'progress': 0,
                    'message': 'Файл загружен, ожидает обработки',
                    'file_path': str(file_path),
                    'metadata': {'model': selected_model}
                }
                
                if not self.create_job_in_db(job_data):
                    flash('Ошибка создания задачи', 'error')
                    return redirect(url_for('index'))
                
                # Запускаем обработку в отдельном потоке
                logger.info(f"🚀 Запуск обработки файла {job_id} в отдельном потоке")
                future = self.executor.submit(self.process_file_sync, job_id)
                logger.info(f"🚀 Задача {job_id} отправлена в ThreadPoolExecutor: {future}")
                
                session['current_job_id'] = job_id
                flash(f'Файл "{filename}" успешно загружен и поставлен в очередь на обработку', 'success')
                return redirect(url_for('status', job_id=job_id))
                
            except RequestEntityTooLarge:
                max_size = self.app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024)
                flash(f'Файл слишком большой. Максимальный размер: {max_size} МБ', 'error')
                return redirect(url_for('index'))
            except Exception as e:
                logger.error(f"❌ Ошибка загрузки файла: {e}")
                flash(f'Ошибка загрузки файла: {str(e)}', 'error')
                return redirect(url_for('index'))
        
        @self.app.route('/status/<job_id>')
        @require_auth()
        def status(job_id: str):
            """Страница статуса обработки"""
            job = self.get_job_status(job_id)
            if not job:
                flash('Задача не найдена или у вас нет доступа к ней', 'error')
                return redirect(url_for('index'))
            
            templates = self.get_available_templates()
            available_models = self.get_available_models()
            current_model = self.get_job_model(job)

            return render_template_string(
                self.templates.get_status_template(),
                job_id=job_id,
                job=job,
                templates=templates,
                available_models=available_models,
                current_model=current_model
            )
        
        @self.app.route('/api/status/<job_id>')
        @require_auth(redirect_on_failure=False)
        def api_status(job_id: str):
            """API для получения статуса задачи"""
            job = self.get_job_status(job_id)
            if not job:
                return jsonify({'error': 'Job not found or access denied'}), 404
            
            return jsonify({
                'status': job['status'],
                'progress': job['progress'],
                'message': job['message'],
                'filename': job['filename'],
                'template': job['template']
            })
        
        @self.app.route('/download/<job_id>/<file_type>')
        @require_auth()
        def download_file(job_id: str, file_type: str):
            """Скачивание результирующих файлов"""
            job = self.get_job_status(job_id)
            if not job:
                flash('Задача не найдена или у вас нет доступа к ней', 'error')
                return redirect(url_for('index'))
            
            if job['status'] != 'completed':
                flash('Обработка еще не завершена', 'error')
                return redirect(url_for('status', job_id=job_id))
            
            try:
                if file_type == 'transcript':
                    file_path = job.get('transcript_file')
                    download_name = f"{Path(job['filename']).stem}_transcript.txt"
                elif file_type == 'summary':
                    file_path = job.get('summary_file')
                    download_name = f"{Path(job['filename']).stem}_summary.md"
                else:
                    flash('Неизвестный тип файла', 'error')
                    return redirect(url_for('status', job_id=job_id))
                
                if not file_path or not os.path.exists(file_path):
                    flash('Файл не найден', 'error')
                    return redirect(url_for('status', job_id=job_id))
                
                return send_file(file_path, as_attachment=True, download_name=download_name)
                
            except Exception as e:
                logger.error(f"❌ Ошибка скачивания файла: {e}")
                flash(f'Ошибка скачивания файла: {str(e)}', 'error')
                return redirect(url_for('status', job_id=job_id))
        
        @self.app.route('/view/<job_id>/<file_type>')
        @require_auth()
        def view_file(job_id: str, file_type: str):
            """Просмотр файлов в веб-интерфейсе"""
            job = self.get_job_status(job_id)
            if not job:
                flash('Задача не найдена или у вас нет доступа к ней', 'error')
                return redirect(url_for('index'))
            
            if job['status'] != 'completed':
                flash('Обработка еще не завершена', 'error')
                return redirect(url_for('status', job_id=job_id))
            
            try:
                if file_type == 'transcript':
                    file_path = job.get('transcript_file')
                    file_title = "Транскрипт встречи"
                    is_markdown = False
                elif file_type == 'summary':
                    file_path = job.get('summary_file')
                    file_title = "Протокол встречи"
                    is_markdown = True
                else:
                    flash('Неизвестный тип файла', 'error')
                    return redirect(url_for('status', job_id=job_id))
                
                if not file_path or not os.path.exists(file_path):
                    flash('Файл не найден', 'error')
                    return redirect(url_for('status', job_id=job_id))
                
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                return render_template_string(
                    self.templates.get_view_template(),
                    content=content,
                    file_title=file_title,
                    filename=job['filename'],
                    job_id=job_id,
                    file_type=file_type,
                    is_markdown=is_markdown
                )
                
            except Exception as e:
                logger.error(f"❌ Ошибка просмотра файла: {e}")
                flash(f'Ошибка просмотра файла: {str(e)}', 'error')
                return redirect(url_for('status', job_id=job_id))
        
        @self.app.route('/generate_protocol/<job_id>', methods=['POST'])
        @require_auth()
        def generate_protocol(job_id: str):
            """Генерация протокола в новом шаблоне из готового транскрипта"""
            job = self.get_job_status(job_id)
            if not job:
                flash('Задача не найдена', 'error')
                return redirect(url_for('index'))
            
            if job['status'] != 'completed':
                flash('Обработка еще не завершена', 'error')
                return redirect(url_for('status', job_id=job_id))
            
            new_template = request.form.get('new_template')
            if not new_template:
                flash('Не выбран шаблон', 'error')
                return redirect(url_for('status', job_id=job_id))

            # Модель: либо явно выбрана в форме, либо берём ту же, что и у исходной задачи
            selected_model = self.resolve_model(
                request.form.get('model') or self.get_job_model(job)
            )

            # Проверяем наличие транскрипта
            transcript_file = job.get('transcript_file')
            if not transcript_file or not os.path.exists(transcript_file):
                flash('Файл транскрипта не найден', 'error')
                return redirect(url_for('status', job_id=job_id))

            try:
                # Создаем новый ID для задачи генерации протокола
                protocol_job_id = f"{job_id}_protocol_{new_template}"
                user_id = get_current_user_id()

                # Создаем задачу генерации протокола в базе данных
                protocol_job_data = {
                    'job_id': protocol_job_id,
                    'user_id': user_id,
                    'filename': f"{job['filename']} (протокол {new_template})",
                    'template': new_template,
                    'status': 'processing',
                    'progress': 0,
                    'message': f'Генерация протокола в шаблоне "{new_template}"...',
                    'transcript_file': transcript_file,
                    'metadata': {'model': selected_model}
                }
                
                if not self.create_job_in_db(protocol_job_data):
                    flash('Ошибка создания задачи генерации протокола', 'error')
                    return redirect(url_for('status', job_id=job_id))
                
                # Запускаем генерацию протокола в отдельном потоке
                self.executor.submit(self.generate_protocol_sync, protocol_job_id, transcript_file, new_template)

                flash(f'Запущена генерация протокола в шаблоне "{new_template}" (модель: {selected_model})', 'success')
                return redirect(url_for('status', job_id=protocol_job_id))
                
            except Exception as e:
                logger.error(f"❌ Ошибка запуска генерации протокола: {e}")
                flash(f'Ошибка запуска генерации протокола: {str(e)}', 'error')
                return redirect(url_for('status', job_id=job_id))
        
        @self.app.route('/jobs')
        @require_auth()
        def jobs_list():
            """Список задач пользователя"""
            try:
                user_id = get_current_user_id()
                if not user_id:
                    flash('Ошибка аутентификации', 'error')
                    return redirect(url_for('index'))
                
                # Получаем задачи пользователя из базы данных
                user_jobs = self.db_manager.get_user_jobs(user_id, limit=50)
                
                jobs = []
                for job_data in user_jobs:
                    created_at = job_data.get('created_at')
                    if isinstance(created_at, str):
                        try:
                            created_at = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                        except:
                            created_at = datetime.utcnow()
                    elif not isinstance(created_at, datetime):
                        created_at = datetime.utcnow()
                    
                    jobs.append({
                        'id': job_data['job_id'],
                        'filename': job_data['filename'],
                        'status': job_data['status'],
                        'template': job_data['template'],
                        'created_at': created_at.strftime('%Y-%m-%d %H:%M:%S'),
                        'progress': job_data.get('progress', 0)
                    })
                
                return render_template_string(self.templates.get_jobs_template(), jobs=jobs)
                
            except Exception as e:
                logger.error(f"Ошибка получения списка задач: {e}")
                flash('Ошибка получения списка задач', 'error')
                return redirect(url_for('index'))
        
        @self.app.route('/statistics')
        @require_auth()
        def statistics():
            """Страница статистики использования приложения"""
            try:
                # Получаем параметры из query string
                days_back = request.args.get('days', type=int)
                start_date = request.args.get('start_date', type=str)
                end_date = request.args.get('end_date', type=str)
                
                # Валидация и обработка параметров
                if start_date and end_date:
                    # Используем диапазон дат
                    # Даты приходят в формате YYYY-MM-DD из HTML5 date input
                    try:
                        # Проверяем формат дат
                        from datetime import datetime
                        datetime.fromisoformat(start_date)
                        datetime.fromisoformat(end_date)
                        
                        # Получаем статистику с диапазоном дат
                        stats = self.db_manager.get_usage_statistics(
                            start_date=start_date,
                            end_date=end_date
                        )
                        
                        # Вычисляем количество дней для отображения
                        start_dt = datetime.fromisoformat(start_date)
                        end_dt = datetime.fromisoformat(end_date)
                        days_back = (end_dt - start_dt).days + 1
                        
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Неверный формат дат: {e}")
                        # Fallback на 30 дней
                        days_back = 30
                        start_date = None
                        end_date = None
                        stats = self.db_manager.get_usage_statistics(days_back=days_back)
                
                elif days_back:
                    # Используем параметр days для обратной совместимости
                    if days_back < 1:
                        days_back = 30
                    elif days_back > 365:
                        days_back = 365
                    
                    stats = self.db_manager.get_usage_statistics(days_back=days_back)
                    start_date = None
                    end_date = None
                
                else:
                    # По умолчанию 30 дней
                    days_back = 30
                    stats = self.db_manager.get_usage_statistics(days_back=days_back)
                    start_date = None
                    end_date = None
                
                return render_template_string(
                    self.templates.get_statistics_template(),
                    stats=stats,
                    days_back=days_back,
                    start_date=start_date,
                    end_date=end_date
                )
                
            except Exception as e:
                logger.error(f"Ошибка получения статистики: {e}")
                flash('Ошибка получения статистики', 'error')
                return redirect(url_for('index'))
        
        @self.app.route('/statistics/export')
        @require_auth()
        def export_statistics():
            """Экспорт сырых данных статистики в Excel"""
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.utils import get_column_letter
                from io import BytesIO
                from datetime import datetime
                
                # Получаем параметры фильтрации
                days_back = request.args.get('days', type=int)
                start_date = request.args.get('start_date', type=str)
                end_date = request.args.get('end_date', type=str)
                
                # Определяем диапазон дат для SQL запроса
                if start_date and end_date:
                    try:
                        datetime.fromisoformat(start_date)
                        datetime.fromisoformat(end_date)
                        date_filter = "j.created_at >= ? AND j.created_at < datetime(?, '+1 day')"
                        date_params = (start_date, end_date)
                        start_dt = datetime.fromisoformat(start_date)
                        end_dt = datetime.fromisoformat(end_date)
                        days_back = (end_dt - start_dt).days + 1
                    except (ValueError, TypeError) as e:
                        logger.warning(f"Неверный формат дат при экспорте: {e}")
                        days_back = 30
                        start_date = None
                        end_date = None
                        date_filter = "j.created_at >= datetime('now', '-30 days')"
                        date_params = ()
                elif days_back:
                    if days_back < 1:
                        days_back = 30
                    elif days_back > 365:
                        days_back = 365
                    date_filter = f"j.created_at >= datetime('now', '-{days_back} days')"
                    date_params = ()
                    start_date = None
                    end_date = None
                else:
                    days_back = 30
                    date_filter = "j.created_at >= datetime('now', '-30 days')"
                    date_params = ()
                    start_date = None
                    end_date = None
                
                # Получаем сырые данные из БД
                with self.db_manager._get_connection() as conn:
                    cursor = conn.cursor()
                    
                    # Запрос для получения всех записей из таблицы jobs за период
                    query = f"""
                        SELECT
                            j.job_id,
                            j.user_id,
                            u.name as user_name,
                            u.email as user_email,
                            j.filename,
                            j.template,
                            j.status,
                            j.progress,
                            j.message,
                            j.created_at,
                            j.completed_at,
                            j.error
                        FROM jobs j
                        LEFT JOIN users u ON j.user_id = u.user_id
                        WHERE {date_filter}
                        ORDER BY j.created_at DESC
                    """
                    
                    cursor.execute(query, date_params)
                    rows = cursor.fetchall()
                
                # Создаем Excel файл
                wb = Workbook()
                ws = wb.active
                ws.title = "Статистика"
                
                # Стили
                header_font = Font(bold=True, color="FFFFFF", size=11)
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                data_alignment = Alignment(horizontal="left", vertical="center")
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # Заголовок с информацией о периоде
                ws['A1'] = "Статистика обработки встреч"
                ws['A1'].font = Font(bold=True, size=14)
                ws.merge_cells('A1:L1')
                
                if start_date and end_date:
                    period_text = f"Период: с {start_date} по {end_date} ({days_back} дней)"
                else:
                    period_text = f"Период: последние {days_back} дней"
                
                ws['A2'] = period_text
                ws['A2'].font = Font(italic=True)
                ws.merge_cells('A2:L2')
                
                ws['A3'] = f"Дата экспорта: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ws['A3'].font = Font(italic=True, size=9)
                ws.merge_cells('A3:L3')
                
                # Заголовки колонок
                headers = [
                    "ID задачи",
                    "ID пользователя",
                    "Имя пользователя",
                    "Email пользователя",
                    "Имя файла",
                    "Шаблон",
                    "Статус",
                    "Прогресс (%)",
                    "Сообщение",
                    "Дата создания",
                    "Дата завершения",
                    "Ошибка"
                ]
                
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=5, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # Данные
                for row_idx, row_data in enumerate(rows, 6):
                    row_dict = dict(row_data)
                    
                    ws.cell(row=row_idx, column=1, value=row_dict['job_id']).border = thin_border
                    ws.cell(row=row_idx, column=2, value=row_dict['user_id']).border = thin_border
                    ws.cell(row=row_idx, column=3, value=row_dict.get('user_name') or '').border = thin_border
                    ws.cell(row=row_idx, column=4, value=row_dict.get('user_email') or '').border = thin_border
                    ws.cell(row=row_idx, column=5, value=row_dict['filename']).border = thin_border
                    ws.cell(row=row_idx, column=6, value=row_dict['template']).border = thin_border
                    ws.cell(row=row_idx, column=7, value=row_dict['status']).border = thin_border
                    ws.cell(row=row_idx, column=8, value=row_dict['progress']).border = thin_border
                    ws.cell(row=row_idx, column=9, value=row_dict.get('message') or '').border = thin_border
                    ws.cell(row=row_idx, column=10, value=row_dict.get('created_at') or '').border = thin_border
                    ws.cell(row=row_idx, column=11, value=row_dict.get('completed_at') or '').border = thin_border
                    ws.cell(row=row_idx, column=12, value=row_dict.get('error') or '').border = thin_border
                    
                    # Выравнивание
                    for col in range(1, 13):
                        ws.cell(row=row_idx, column=col).alignment = data_alignment
                
                # Автоширина колонок
                column_widths = {
                    'A': 38,  # ID задачи
                    'B': 25,  # ID пользователя
                    'C': 25,  # Имя пользователя
                    'D': 30,  # Email
                    'E': 40,  # Имя файла
                    'F': 15,  # Шаблон
                    'G': 12,  # Статус
                    'H': 12,  # Прогресс
                    'I': 50,  # Сообщение
                    'J': 20,  # Дата создания
                    'K': 20,  # Дата завершения
                    'L': 50   # Ошибка
                }
                
                for col_letter, width in column_widths.items():
                    ws.column_dimensions[col_letter].width = width
                
                # Закрепляем заголовки
                ws.freeze_panes = 'A6'
                
                # Сохраняем в BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)
                
                # Формируем имя файла
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                if start_date and end_date:
                    filename = f"statistics_raw_{start_date}_{end_date}_{timestamp}.xlsx"
                else:
                    filename = f"statistics_raw_{days_back}days_{timestamp}.xlsx"
                
                logger.info(f"Экспорт сырых данных статистики в Excel: {filename}, записей: {len(rows)} (пользователь: {get_current_user_id()})")
                
                return send_file(
                    output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    download_name=filename
                )
                
            except ImportError:
                logger.error("Библиотека openpyxl не установлена")
                flash('Ошибка: библиотека для экспорта в Excel не установлена', 'error')
                return redirect(url_for('statistics'))
            except Exception as e:
                logger.error(f"Ошибка экспорта статистики в Excel: {e}")
                import traceback
                traceback.print_exc()
                flash(f'Ошибка экспорта статистики: {str(e)}', 'error')
                return redirect(url_for('statistics'))
        
        @self.app.route('/docs')
        def docs_index():
            """Главная страница документации"""
            return render_template_string(self.templates.get_docs_index_template())
        
        @self.app.route('/docs/<doc_name>')
        def view_docs(doc_name: str):
            """Просмотр конкретного документа"""
            docs_map = {
                'guidelines': 'meeting_recording_guidelines.md',
                'checklist': 'quick_meeting_checklist.md',
                'setup': 'recording_setup_guide.md'
            }
            
            if doc_name not in docs_map:
                flash('Документ не найден', 'error')
                return redirect(url_for('docs_index'))
            
            try:
                file_path = docs_map[doc_name]
                if not os.path.exists(file_path):
                    flash('Файл документации не найден', 'error')
                    return redirect(url_for('docs_index'))
                
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                
                doc_titles = {
                    'guidelines': 'Полное руководство по проведению встреч',
                    'checklist': 'Быстрый чек-лист для записи встреч',
                    'setup': 'Техническое руководство по настройке записи'
                }
                
                return render_template_string(
                    self.templates.get_docs_view_template(),
                    content=content,
                    doc_title=doc_titles[doc_name],
                    doc_name=doc_name
                )
                
            except Exception as e:
                logger.error(f"❌ Ошибка чтения документации: {e}")
                flash(f'Ошибка чтения документации: {str(e)}', 'error')
                return redirect(url_for('docs_index'))
        
        @self.app.route('/publish_confluence/<job_id>', methods=['POST'])
        @require_auth()
        def publish_confluence(job_id: str):
            """Публикация протокола в Confluence"""
            try:
                # Проверяем существование задачи и права доступа
                job = self.get_job_status(job_id)
                if not job:
                    return jsonify({'success': False, 'error': 'Задача не найдена или у вас нет доступа к ней'}), 404
                
                if job['status'] != 'completed':
                    return jsonify({'success': False, 'error': 'Обработка еще не завершена'}), 400
                
                # Получаем данные формы
                base_page_url = request.form.get('base_page_url', '').strip()
                page_title = request.form.get('page_title', '').strip()
                
                logger.info(f"🔍 DEBUG: Confluence publication form data:")
                logger.info(f"   base_page_url: {base_page_url}")
                logger.info(f"   page_title: {page_title}")
                
                if not base_page_url:
                    return jsonify({'success': False, 'error': 'URL базовой страницы обязателен'}), 400
                
                # Получаем настройки Confluence из конфигурации (нужно для поиска страниц)
                confluence_config = self.config.get('confluence', {})
                if not confluence_config:
                    return jsonify({'success': False, 'error': 'Confluence не настроен в конфигурации'}), 500
                
                # Валидация URL - поддерживаем только Confluence Server форматы
                import re
                
                # Confluence Server формат 1: https://wiki.domain.com/pages/viewpage.action?pageId=123456
                server_pattern1 = r'^https?://[^/]+/pages/viewpage\.action\?pageId=\d+'
                
                # Confluence Server формат 2: https://wiki.domain.com/display/SPACE/PAGE
                server_pattern2 = r'^https?://[^/]+/display/[^/]+/[^/]+'
                
                is_server1 = re.match(server_pattern1, base_page_url)
                is_server2 = re.match(server_pattern2, base_page_url)
                
                if not (is_server1 or is_server2):
                    return jsonify({
                        'success': False,
                        'error': 'Неверный формат URL Confluence Server. Поддерживаются форматы:\n' +
                                '• Server: https://wiki.domain.com/pages/viewpage.action?pageId=123456\n' +
                                '• Server: https://wiki.domain.com/display/SPACE/PAGE'
                    }), 400
                
                # Инициализируем parent_page_id сразу
                parent_page_id = None
                
                # Используем новую улучшенную функцию извлечения метаданных
                logger.info(f"🔍 DEBUG: Извлечение метаданных из URL: {base_page_url}")
                
                # Получаем API токен для аутентификации - сначала из переменных окружения, потом из конфига
                api_token = os.getenv('CONFLUENCE_API_TOKEN') or confluence_config.get('api_token', '')
                logger.info(f"🔍 DEBUG: API токен {'найден' if api_token else 'НЕ НАЙДЕН'}")
                
                # Извлекаем метаданные с помощью улучшенной функции
                metadata = self.extract_confluence_metadata(base_page_url, api_token, timeout=30)
                
                # Используем извлеченные данные
                extracted_page_id = metadata.get('page_id')
                extracted_space_key = metadata.get('space_key')
                extracted_page_title = metadata.get('page_title')
                
                # Используем space_key только из автоматического извлечения
                space_key = extracted_space_key
                if space_key:
                    logger.info(f"🔍 DEBUG: Используем space_key из метаданных: {space_key}")
                else:
                    logger.warning(f"🔍 DEBUG: space_key не удалось извлечь из метаданных")
                
                # Используем извлеченный page_id как parent_page_id
                parent_page_id = extracted_page_id
                
                # Проверяем наличие файла протокола
                summary_file = job.get('summary_file')
                if not summary_file or not os.path.exists(summary_file):
                    return jsonify({'success': False, 'error': 'Файл протокола не найден'}), 404
                
                # Читаем содержимое протокола
                with open(summary_file, 'r', encoding='utf-8') as f:
                    protocol_content = f.read()
                
                # ИСПРАВЛЕНИЕ: Автогенерация заголовка ВСЕГДА происходит первой (если заголовок не указан пользователем)
                if not page_title:
                    from confluence_client import ConfluenceContentProcessor
                    processor = ConfluenceContentProcessor()
                    
                    # Извлекаем информацию о встрече из протокола
                    meeting_date, meeting_topic = processor.extract_meeting_info(protocol_content)
                    
                    # Генерируем заголовок по правильному шаблону YYYYMMDD - <тема встречи>
                    page_title = processor.generate_page_title(
                        meeting_date,
                        meeting_topic,
                        job.get('filename')
                    )
                    logger.info(f"🔍 DEBUG: Автогенерированный заголовок: {page_title}")
                
                # Fallback: Если автогенерация не сработала и есть заголовок из метаданных, используем его
                if not page_title and extracted_page_title:
                    page_title = extracted_page_title
                    logger.info(f"🔍 DEBUG: Fallback - используем заголовок из метаданных: {page_title}")
                
                logger.info(f"🔍 DEBUG: Итоговые данные:")
                logger.info(f"   parent_page_id: {parent_page_id}")
                logger.info(f"   space_key: {space_key}")
                logger.info(f"   page_title: {page_title}")
                logger.info(f"   extraction_method: {metadata.get('extraction_method')}")
                
                # Импортируем и инициализируем Confluence клиент
                try:
                    from confluence_client import ConfluenceServerClient, ConfluenceConfig
                    
                    # Получаем настройки Confluence из конфигурации
                    confluence_config = self.config.get('confluence', {})
                    if not confluence_config:
                        return jsonify({'success': False, 'error': 'Confluence не настроен в конфигурации'}), 500
                    
                    # Получаем API токен из переменных окружения или конфигурации
                    api_token = os.getenv('CONFLUENCE_API_TOKEN') or confluence_config.get('api_token', '')
                    
                    if not api_token:
                        return jsonify({'success': False, 'error': 'Не указан CONFLUENCE_API_TOKEN в переменных окружения или конфигурации'}), 500
                    
                    # Создаем конфигурацию
                    config = ConfluenceConfig(
                        base_url=confluence_config['base_url'],
                        api_token=api_token,
                        space_key=space_key,
                        username=confluence_config.get('username'),  # Опциональное поле для совместимости
                        timeout=confluence_config.get('timeout', 30),
                        max_retries=confluence_config.get('max_retries', 3),
                        retry_delay=confluence_config.get('retry_delay', 1.0)
                    )
                    
                    confluence_client = ConfluenceServerClient(config)
                    
                    # Конвертируем markdown в Confluence Storage Format
                    from confluence_client import ConfluenceContentProcessor
                    processor = ConfluenceContentProcessor()
                    confluence_content = processor.markdown_to_confluence(protocol_content)
                    
                    # Получаем метки из конфигурации
                    page_labels = confluence_config.get('advanced_settings', {}).get('page_labels', [])
                    logger.info(f"🔍 DEBUG: Метки из конфигурации: {page_labels}")
                    
                    # Создаем страницу в Confluence
                    page_info = confluence_client.create_page(
                        title=page_title,
                        content=confluence_content,
                        parent_page_id=parent_page_id,
                        space_key=space_key,
                        labels=page_labels
                    )
                    
                    logger.info(f"🔍 DEBUG: Created Confluence page with parent_page_id: {parent_page_id}")
                    
                    # Строим правильный URL страницы для Confluence Server
                    page_url = f"{confluence_config['base_url'].rstrip('/')}/pages/viewpage.action?pageId={page_info['id']}"
                    
                    # Сохраняем информацию о публикации в базе данных
                    user_id = get_current_user_id()
                    publication_data = {
                        'job_id': job_id,
                        'confluence_page_id': page_info['id'],
                        'confluence_page_url': page_url,
                        'confluence_space_key': space_key,
                        'page_title': page_title,
                        'publication_status': 'published'
                    }
                    
                    try:
                        self.db_manager.create_confluence_publication(publication_data)
                        logger.info(f"Информация о публикации сохранена в БД: {page_info['id']}")
                    except Exception as db_error:
                        logger.warning(f"Не удалось сохранить информацию о публикации в БД: {db_error}")
                    
                    # Проверяем, это AJAX запрос или обычный
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        # AJAX запрос - возвращаем JSON
                        return jsonify({
                            'success': True,
                            'message': 'Протокол успешно опубликован в Confluence!',
                            'page_url': page_url,
                            'page_id': page_info['id']
                        })
                    else:
                        # Обычный запрос - используем flash и redirect на страницу статуса
                        flash(f'Протокол успешно опубликован в Confluence! <a href="{page_url}" target="_blank">Открыть страницу</a>', 'success')
                        return redirect(url_for('status', job_id=job_id))
                        
                except ImportError:
                    return jsonify({'success': False, 'error': 'Модуль Confluence не найден'}), 500
                except Exception as confluence_error:
                    logger.error(f"Ошибка Confluence клиента: {confluence_error}")
                    
                    # Сохраняем информацию об ошибке
                    user_id = get_current_user_id()
                    publication_data = {
                        'job_id': job_id,
                        'confluence_page_id': '',  # Пустой ID при ошибке
                        'confluence_page_url': base_page_url,  # Используем исходный URL
                        'confluence_space_key': space_key,
                        'page_title': page_title,
                        'publication_status': 'failed',
                        'error_message': str(confluence_error)
                    }
                    
                    try:
                        self.db_manager.create_confluence_publication(publication_data)
                        logger.info(f"Информация об ошибке публикации сохранена в БД")
                    except Exception as db_error:
                        logger.warning(f"Не удалось сохранить информацию об ошибке в БД: {db_error}")
                    
                    # Проверяем, это AJAX запрос или обычный
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        # AJAX запрос - возвращаем JSON
                        return jsonify({
                            'success': False,
                            'error': f'Ошибка публикации в Confluence: {str(confluence_error)}'
                        })
                    else:
                        # Обычный запрос - используем flash и redirect на страницу статуса
                        flash(f'Ошибка публикации в Confluence: {str(confluence_error)}', 'error')
                        return redirect(url_for('status', job_id=job_id))
                
            except Exception as e:
                logger.error(f"❌ Ошибка публикации в Confluence: {e}")
                
                # Проверяем, это AJAX запрос или обычный
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # AJAX запрос - возвращаем JSON
                    return jsonify({
                        'success': False,
                        'error': f'Внутренняя ошибка сервера: {str(e)}'
                    })
                else:
                    # Обычный запрос - используем flash и redirect на страницу статуса
                    flash(f'Внутренняя ошибка сервера: {str(e)}', 'error')
                    return redirect(url_for('status', job_id=job_id))
        
        @self.app.route('/confluence_publications/<job_id>')
        @require_auth()
        def confluence_publications(job_id: str):
            """Получение истории публикаций Confluence для задачи"""
            try:
                # Проверяем существование задачи и права доступа
                job = self.get_job_status(job_id)
                if not job:
                    return jsonify({'error': 'Задача не найдена или у вас нет доступа к ней'}), 404
                
                # Получаем историю публикаций из базы данных
                user_id = get_current_user_id()
                publications = self.db_manager.get_confluence_publications(job_id, user_id)
                
                logger.info(f"🔍 DEBUG: Publication history request for job {job_id}, user {user_id}")
                logger.info(f"🔍 DEBUG: Found {len(publications)} publications")
                for i, pub in enumerate(publications):
                    logger.info(f"🔍 DEBUG: Publication {i+1}: {pub}")
                
                return jsonify({
                    'publications': publications,
                    'count': len(publications)
                })
                
            except Exception as e:
                logger.error(f"❌ Ошибка получения истории публикаций: {e}")
                return jsonify({'error': f'Ошибка получения истории: {str(e)}'}), 500
        
        @self.app.errorhandler(RequestEntityTooLarge)
        def handle_file_too_large(e):
            max_size = self.app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024)
            flash(f'Файл слишком большой. Максимальный размер: {max_size} МБ', 'error')
            return redirect(url_for('index'))
    
    def process_file_sync(self, job_id: str):
        """Синхронная обработка файла в отдельном потоке"""
        logger.info(f"🔄 process_file_sync запущена для задачи {job_id}")
        
        # Получаем задачу из базы данных
        job = None
        try:
            logger.debug(f"🔍 Получение задачи {job_id} из базы данных")
            # Получаем задачу без проверки пользователя (для фонового процесса)
            with self.db_manager._get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM jobs WHERE job_id = ?", (job_id,))
                row = cursor.fetchone()
                if row:
                    job = dict(row)
                    logger.info(f"✅ Задача {job_id} найдена в базе данных: {job['filename']}")
                else:
                    logger.error(f"❌ Задача {job_id} не найдена в базе данных")
        except Exception as e:
            logger.error(f"❌ Ошибка получения задачи {job_id}: {e}")
            return
        
        if not job:
            logger.error(f"❌ Задача {job_id} не найдена")
            return
        
        def progress_callback(progress: int, message: str):
            """Callback для обновления прогресса"""
            try:
                with self.db_manager._get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        "UPDATE jobs SET progress = ?, message = ? WHERE job_id = ?",
                        (progress, message, job_id)
                    )
                    conn.commit()
            except Exception as e:
                logger.error(f"Ошибка обновления прогресса для {job_id}: {e}")
        
        try:
            logger.info(f"🔄 Начало обработки файла {job_id}: {job['filename']} для пользователя {job['user_id']}")
            
            # Проверяем существование файла
            input_file_path = job['file_path']
            if not os.path.exists(input_file_path):
                raise Exception(f"Файл не найден: {input_file_path}")
            
            logger.info(f"📂 Входной файл: {input_file_path}")
            logger.info(f"📊 Размер файла: {os.path.getsize(input_file_path)} байт")
            
            # Создаем пользовательскую директорию для вывода
            user_output_dir = self.get_user_output_dir(job['user_id'])
            output_dir = user_output_dir / job_id
            output_dir.mkdir(parents=True, exist_ok=True)
            
            job_model = self.get_job_model(job)
            logger.info(f"🤖 Модель для задачи {job_id}: {job_model}")

            processor = MeetingProcessor(
                deepgram_api_key=self.deepgram_key,
                claude_api_key=self.claude_key,
                deepgram_timeout=self.processing_settings.get('deepgram_timeout_seconds', 300),
                claude_model=job_model,
                chunk_duration_minutes=self.processing_settings.get('chunk_duration_minutes', 15),
                template_type=job['template'],
                progress_callback=progress_callback,
                deepgram_language=self.processing_settings.get('language', 'ru'),
                deepgram_model=self.processing_settings.get('deepgram_model', 'nova-2')
            )
            
            # Основная обработка
            success = processor.process_meeting(
                input_file_path=job['file_path'],
                output_dir=str(output_dir),
                name_mapping=None,
                keep_audio_file=False,
                template_type=job['template']
            )
            
            if success:
                # Извлекаем исходное имя файла
                original_filename = Path(job['file_path']).name
                if original_filename.startswith(job_id + '_'):
                    original_filename = original_filename[len(job_id) + 1:]
                
                input_name = Path(original_filename).stem
                transcript_file = output_dir / f"{input_name}_transcript.txt"
                summary_file = output_dir / f"{input_name}_summary.md"
                
                if transcript_file.exists() and summary_file.exists():
                    # Обновляем задачу в базе данных
                    with self.db_manager._get_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute("""
                            UPDATE jobs SET
                                status = 'completed',
                                progress = 100,
                                message = 'Обработка завершена успешно!',
                                transcript_file = ?,
                                summary_file = ?,
                                completed_at = CURRENT_TIMESTAMP
                            WHERE job_id = ?
                        """, (str(transcript_file), str(summary_file), job_id))
                        conn.commit()
                    
                    logger.info(f"✅ Обработка файла {job_id} завершена успешно")
                else:
                    # Ищем любые файлы в директории
                    all_files = list(output_dir.glob("*"))
                    transcript_files = [f for f in all_files if "_transcript.txt" in f.name]
                    summary_files = [f for f in all_files if "_summary.md" in f.name]
                    
                    logger.info(f"📁 Все файлы в {output_dir}: {[f.name for f in all_files]}")
                    
                    if transcript_files and summary_files:
                        transcript_file = transcript_files[0]
                        summary_file = summary_files[0]
                        
                        # Обновляем задачу в базе данных
                        with self.db_manager._get_connection() as conn:
                            cursor = conn.cursor()
                            cursor.execute("""
                                UPDATE jobs SET
                                    status = 'completed',
                                    progress = 100,
                                    message = 'Обработка завершена успешно!',
                                    transcript_file = ?,
                                    summary_file = ?,
                                    completed_at = CURRENT_TIMESTAMP
                                WHERE job_id = ?
                            """, (str(transcript_file), str(summary_file), job_id))
                            conn.commit()
                        
                        logger.info(f"✅ Обработка файла {job_id} завершена успешно")
                        return
                    
                    raise Exception(f"Файлы не найдены. Есть: {[f.name for f in all_files]}")
            else:
                raise Exception("Обработка завершилась с ошибкой")
                
        except Exception as e:
            logger.error(f"❌ Ошибка обработки файла {job_id}: {e}")
            # Обновляем статус ошибки в базе данных
            try:
                with self.db_manager._get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        UPDATE jobs SET
                            status = 'error',
                            progress = 0,
                            message = ?,
                            error = ?
                        WHERE job_id = ?
                    """, (f'Ошибка обработки: {str(e)}', str(e), job_id))
                    conn.commit()
            except Exception as db_error:
                logger.error(f"Ошибка обновления статуса ошибки в БД: {db_error}")
        
        finally:
            # Очищаем исходный файл
            try:
                if job and job.get('file_path') and os.path.exists(job['file_path']):
                    os.remove(job['file_path'])
                    logger.debug(f"Удален временный файл: {job['file_path']}")
            except Exception as e:
                logger.warning(f"Не удалось удалить временный файл: {e}")
    
    def generate_protocol_sync(self, job_id: str, transcript_file: str, template_type: str):
        """Синхронная генерация протокола из транскрипта в отдельном потоке"""
        logger.info(f"🔄 generate_protocol_sync запущена для задачи {job_id}")
        
        # Получаем задачу из базы данных напрямую (без проверки аутентификации для фонового процесса)
        job = None
        try:
            logger.debug(f"🔍 Получение задачи {job_id} из базы данных")
            with self.db_manager._get_connection() as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM jobs WHERE job_id = ?", (job_id,))
                row = cursor.fetchone()
                if row:
                    job = dict(row)
                    logger.info(f"✅ Задача {job_id} найдена в базе данных: {job['filename']}")
                else:
                    logger.error(f"❌ Задача {job_id} не найдена в базе данных")
                    return
        except Exception as e:
            logger.error(f"❌ Ошибка получения задачи {job_id}: {e}")
            return
        
        def progress_callback(progress: int, message: str):
            """Callback для обновления прогресса"""
            try:
                with self.db_manager._get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute(
                        "UPDATE jobs SET progress = ?, message = ? WHERE job_id = ?",
                        (progress, message, job_id)
                    )
                    conn.commit()
                    logger.debug(f"📊 Прогресс обновлен для {job_id}: {progress}% - {message}")
            except Exception as e:
                logger.error(f"Ошибка обновления прогресса для {job_id}: {e}")
        
        try:
            # Создаем выходную директорию для нового протокола
            output_dir = self.output_folder / job_id
            output_dir.mkdir(exist_ok=True)
            
            # Создаем процессор только для генерации протокола
            job_model = self.get_job_model(job)
            logger.info(f"🤖 Создание MeetingProcessor для генерации протокола")
            logger.info(f"   Claude API ключ: {'✅ установлен' if self.claude_key else '❌ отсутствует'}")
            logger.info(f"   Модель: {job_model}")
            logger.info(f"   Тип шаблона: {template_type}")
            logger.info(f"   Файл транскрипта: {transcript_file}")
            logger.info(f"   Выходная директория: {output_dir}")

            processor = MeetingProcessor(
                deepgram_api_key="dummy",  # Не нужен для генерации протокола
                claude_api_key=self.claude_key,
                claude_model=job_model,
                template_type=template_type,
                templates_config_file=self.config.get("paths", {}).get("templates_config", "templates_config.json"),
                team_config_file=self.config.get("paths", {}).get("team_config", "team_config.json"),
                progress_callback=progress_callback
            )
            
            logger.info(f"🚀 Запуск генерации протокола из транскрипта")
            
            # Генерируем протокол из транскрипта
            success = processor.generate_protocol_from_transcript(
                transcript_file_path=transcript_file,
                output_dir=str(output_dir),
                template_type=template_type
            )
            
            logger.info(f"📊 Результат генерации протокола: {'✅ успех' if success else '❌ ошибка'}")
            
            if success:
                # Ищем сгенерированный протокол
                all_files = list(output_dir.glob("*"))
                summary_files = [f for f in all_files if "_summary.md" in f.name]
                
                if summary_files:
                    summary_file = summary_files[0]
                    
                    # Обновляем задачу в базе данных напрямую
                    try:
                        with self.db_manager._get_connection() as conn:
                            cursor = conn.cursor()
                            cursor.execute("""
                                UPDATE jobs SET
                                    status = 'completed',
                                    progress = 100,
                                    message = 'Протокол успешно сгенерирован!',
                                    transcript_file = ?,
                                    summary_file = ?,
                                    completed_at = CURRENT_TIMESTAMP
                                WHERE job_id = ?
                            """, (transcript_file, str(summary_file), job_id))
                            conn.commit()
                        
                        logger.info(f"✅ Генерация протокола {job_id} завершена успешно")
                    except Exception as db_error:
                        logger.error(f"Ошибка обновления статуса завершения в БД: {db_error}")
                else:
                    raise Exception(f"Файл протокола не найден. Есть: {[f.name for f in all_files]}")
            else:
                raise Exception("Генерация протокола завершилась с ошибкой")
                
        except Exception as e:
            logger.error(f"❌ Ошибка генерации протокола {job_id}: {e}")
            # Обновляем статус ошибки в базе данных напрямую
            try:
                with self.db_manager._get_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute("""
                        UPDATE jobs SET
                            status = 'error',
                            progress = 0,
                            message = ?,
                            error = ?
                        WHERE job_id = ?
                    """, (f'Ошибка генерации протокола: {str(e)}', str(e), job_id))
                    conn.commit()
            except Exception as db_error:
                logger.error(f"Ошибка обновления статуса ошибки в БД: {db_error}")
    
    def run(self, host: str = '127.0.0.1', port: int = 5000, debug: bool = False):
        """Запуск веб-приложения"""
        logger.info(f"🌐 Запуск веб-приложения на http://{host}:{port}")
        
        try:
            self.app.run(host=host, port=port, debug=debug, threaded=True)
        finally:
            # Закрываем пул потоков при завершении
            self.executor.shutdown(wait=True)

def main():
    """Основная функция"""
    import argparse
    
    parser = argparse.ArgumentParser(description="Рабочий веб-интерфейс для системы обработки встреч")
    parser.add_argument("-c", "--config", default="config.json", help="Путь к файлу конфигурации")
    parser.add_argument("--host", default="127.0.0.1", help="IP адрес для запуска сервера")
    parser.add_argument("--port", type=int, default=5000, help="Порт для запуска сервера")
    parser.add_argument("--debug", action="store_true", help="Запуск в режиме отладки")
    parser.add_argument("--debug-auth", action="store_true", help="Отключить проверку токенов аутентификации (только для разработки)")
    
    args = parser.parse_args()
    
    try:
        # Создаем необходимые директории
        os.makedirs("logs", exist_ok=True)
        os.makedirs("web_uploads", exist_ok=True)
        os.makedirs("web_output", exist_ok=True)
        
        # Проверяем отладочный режим аутентификации
        from config_loader import ConfigLoader
        config = ConfigLoader.load_config(args.config)
        debug_mode_enabled = False
        
        if config and config.get('auth', {}).get('debug_mode', False):
            debug_mode_enabled = True
            print("🔧 ВНИМАНИЕ: Отладочный режим аутентификации включен в конфигурации!")
            print("   Проверка токенов отключена. Используйте только для разработки!")
        
        # Обрабатываем флаг отладочного режима аутентификации
        if args.debug_auth:
            debug_mode_enabled = True
            print("🔧 ВНИМАНИЕ: Отладочный режим аутентификации включен флагом --debug-auth!")
            print("   Проверка токенов отключена. Используйте только для разработки!")
            
            # Загружаем конфигурацию и принудительно включаем отладочный режим
            if config:
                config.setdefault('auth', {})['debug_mode'] = True
                # Сохраняем временно измененную конфигурацию
                import tempfile
                import json
                with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False, encoding='utf-8') as f:
                    json.dump(config, f, indent=2, ensure_ascii=False)
                    temp_config_path = f.name
                args.config = temp_config_path
                print(f"🔧 Создан временный файл конфигурации: {temp_config_path}")
        
        # Создаем веб-приложение
        web_app = WorkingMeetingWebApp(args.config)
        
        print("\n" + "="*60)
        print("🚀 MEETING PROCESSOR WEB SERVER (РАБОЧАЯ ВЕРСИЯ)")
        print("="*60)
        print(f"📱 Веб-интерфейс: http://localhost:{args.port}")
        print("🔧 Конфигурация: config.json")
        print("🔑 API ключи: переменные окружения")
        print("📁 Загрузки: web_uploads/")
        print("📄 Результаты: web_output/")
        print("📊 Логи: logs/web_app.log")
        print("🧵 Многопоточная обработка: ✅")
        print("🔄 Автоочистка файлов: ✅")
        
        # Показываем статус аутентификации
        if args.debug_auth:
            print("🔧 Отладочный режим аутентификации: ✅ (ТОЛЬКО ДЛЯ РАЗРАБОТКИ)")
            print("📊 Статистика доступна: http://localhost:{}/statistics".format(args.port))
        else:
            print("🔐 Аутентификация: ✅ (требуется X-Identity-Token)")
        
        print("\n💡 Для остановки нажмите Ctrl+C")
        print("="*60)
        
        # Запускаем сервер
        web_app.run(host=args.host, port=args.port, debug=args.debug)
        
    except KeyboardInterrupt:
        print("\n👋 Веб-приложение остановлено")
    except Exception as e:
        print(f"❌ Ошибка запуска веб-приложения: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
